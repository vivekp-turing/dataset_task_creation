#!/usr/bin/env python3
"""Select high-quality SEED REPOS (not tasks) from a Turing/SWE-Bench long-range
metadata spreadsheet, filtered against the task requirements.

The output is a list of distinct repositories (deduped, one row per repo) that are
good candidates for authoring ORIGINAL tasks. It is NOT a list of tasks to reuse.

Usage examples:
  # 15 repos each for the top-4 task-spec language groups -> 60 total
  python select_seed_repos.py --xlsx "<sheet>.xlsx" --out seed_repos.csv \
      --per-language 15 --languages "JS/TS,Python,Java,C#"

  # Custom counts per language (overrides --per-language)
  python select_seed_repos.py --xlsx "<sheet>.xlsx" --out seed.csv \
      --lang-counts "JS/TS=20,Python=15,Java=10,Go=5"

Language groups: a group name maps to one or more raw `language` values in the
sheet. JS/TS expands to JavaScript+TypeScript by default. Any sheet language can
be used directly (e.g. Go, Rust, Ruby, C++, C#, PHP).
"""
import argparse
import csv
import sys
from collections import Counter

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")

# Group name -> set of raw `language` values in the sheet.
DEFAULT_GROUP_ALIASES = {
    "JS/TS": {"JavaScript", "TypeScript"},
    "JS": {"JavaScript"},
    "TS": {"TypeScript"},
    "C/C++": {"C", "C++"},
    "C++": {"C++"},
    "C#": {"C#"},
    "Python": {"Python"},
    "Java": {"Java"},
    "Go": {"Go"},
    "Rust": {"Rust"},
    "Ruby": {"Ruby"},
    "PHP": {"PHP"},
    "Swift": {"Swift"},
}

SENTINEL_LOC = 999999  # repo `loc` sentinel meaning unknown / very large


def resolve_group(name):
    """Map a group name to a set of raw sheet language values."""
    if name in DEFAULT_GROUP_ALIASES:
        return DEFAULT_GROUP_ALIASES[name]
    # Fallback: treat the name itself as a raw language value.
    return {name}


def num(v):
    return v if isinstance(v, (int, float)) else None


def load_rows(xlsx_path, sheet_name):
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name is None:
        # Pick the sheet with `instance_id`+`language` headers AND the most rows
        # (several sheets reuse these headers; the raw data sheet is the biggest).
        best = None  # (row_count, title)
        for ws in wb.worksheets:
            first = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if "instance_id" in first and "language" in first:
                cand = (ws.max_row or 0, ws.title)
                if best is None or cand[0] > best[0]:
                    best = cand
        if best is None:
            sys.exit("Could not auto-detect the raw data sheet; pass --sheet.")
        sheet_name = best[1]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    hdr = list(rows[0])
    data = [dict(zip(hdr, r)) for r in rows[1:]]
    return sheet_name, data


def quality_ok(d, args):
    il = num(d.get("instance_loc"))
    lo = num(d.get("loc"))
    st = num(d.get("stars"))
    f2p = num(d.get("f2p_count"))
    if il is None or not (args.min_instance_loc <= il <= args.max_instance_loc):
        return False
    if f2p is None or f2p < args.min_f2p:
        return False
    if st is None or st < args.min_stars:
        return False
    if not args.allow_loc_sentinel and (lo is None or lo >= SENTINEL_LOC):
        return False
    if lo is not None and lo > args.max_repo_loc:
        return False
    if d.get("code_type_primary") not in {"feature", "bug-fix", "refactor"}:
        return False
    return True


def score(d, args):
    s = 0.0
    il = num(d.get("instance_loc")) or 0
    s += max(0.0, 100 - abs(il - args.target_instance_loc) / 3)
    s += (num(d.get("difficulty_score")) or 0) * 15  # favor harder tasks
    s += min(num(d.get("stars")) or 0, 20000) / 1000
    s += min(num(d.get("f2p_count")) or 0, 25)
    ct = d.get("code_type_primary")
    if ct == "bug-fix":
        s += args.bugfix_bonus
    elif ct == "refactor":
        s -= 4
    if (num(d.get("loc")) or 0) < 800:
        s -= 10  # avoid trivial/toy repos
    return s


def parse_lang_counts(args):
    """Return ordered list of (group_name, count)."""
    if args.lang_counts:
        pairs = []
        for chunk in args.lang_counts.split(","):
            name, _, cnt = chunk.partition("=")
            pairs.append((name.strip(), int(cnt)))
        return pairs
    groups = [g.strip() for g in args.languages.split(",") if g.strip()]
    return [(g, args.per_language) for g in groups]


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", required=True, help="Path to the metadata .xlsx")
    ap.add_argument("--out", required=True, help="Output CSV path")
    ap.add_argument("--sheet", default=None,
                    help="Sheet name (auto-detected if omitted)")
    # Language coverage (high-level inputs)
    ap.add_argument("--languages", default="JS/TS,Python,Java,C#",
                    help="Comma-separated group names (used with --per-language)")
    ap.add_argument("--per-language", type=int, default=15,
                    help="Repos to pick per language group")
    ap.add_argument("--lang-counts", default=None,
                    help="Per-group counts, e.g. 'JS/TS=20,Python=15'. Overrides --languages/--per-language")
    # Quality filters (all tunable; defaults match the task spec)
    ap.add_argument("--min-instance-loc", type=int, default=60)
    ap.add_argument("--max-instance-loc", type=int, default=350)
    ap.add_argument("--target-instance-loc", type=int, default=160,
                    help="Sweet spot for ~100 LoC multi-file patches")
    ap.add_argument("--min-f2p", type=int, default=3,
                    help="Minimum fail2pass test count")
    ap.add_argument("--min-stars", type=int, default=250)
    ap.add_argument("--max-repo-loc", type=int, default=2_000_000)
    ap.add_argument("--allow-loc-sentinel", action="store_true",
                    help="Keep repos whose loc is the 999999 unknown/huge sentinel")
    ap.add_argument("--bugfix-bonus", type=float, default=12.0,
                    help="Score bonus to keep a healthy bug-fix mix")
    args = ap.parse_args()

    sheet_name, data = load_rows(args.xlsx, args.sheet)
    lang_plan = parse_lang_counts(args)

    out_rows = []
    summary = []
    for gname, want in lang_plan:
        langs = resolve_group(gname)
        cands = [d for d in data if d.get("language") in langs and quality_ok(d, args)]
        cands.sort(key=lambda d: score(d, args), reverse=True)
        picked, seen = [], set()
        for d in cands:  # one instance per repo -> distinct seed repos
            repo = d.get("repo")
            if repo in seen:
                continue
            picked.append(d)
            seen.add(repo)
            if len(picked) == want:
                break
        ct = Counter(d.get("code_type_primary") for d in picked)
        summary.append((gname, len(cands), len(picked), want, dict(ct)))
        for d in picked:
            out_rows.append({
                "lang_group": gname,
                "repo": d.get("repo"),
                "repo_full_name": d.get("repo_full_name"),
                "language": d.get("language"),
                "instance_loc": num(d.get("instance_loc")),
                "repo_loc": num(d.get("loc")),
                "stars": num(d.get("stars")),
                "f2p_count": num(d.get("f2p_count")),
                "p2p_count": num(d.get("p2p_count")),
                "difficulty_score": d.get("difficulty_score"),
                "code_type_primary": d.get("code_type_primary"),
                "issue_type_primary": d.get("issue_type_primary"),
                "repo_type_primary": d.get("repo_type_primary"),
                "long_horizon_tag": d.get("long_horizon_tag"),
                # Reference only - DO NOT reuse these as tasks; build originals.
                "example_instance_id": d.get("instance_id"),
                "example_pr_url": d.get("pr_url"),
            })

    if not out_rows:
        sys.exit("No repos matched the filters. Loosen thresholds and retry.")

    with open(args.out, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=list(out_rows[0].keys()))
        w.writeheader()
        w.writerows(out_rows)

    print(f"Sheet: {sheet_name}  |  rows analyzed: {len(data)}")
    for gname, elig, got, want, ct in summary:
        flag = "" if got == want else "  <-- SHORTFALL (loosen filters)"
        print(f"  {gname:10s} eligible={elig:>4d}  picked={got}/{want}  mix={ct}{flag}")
    print(f"Wrote {len(out_rows)} distinct seed repos -> {args.out}")


if __name__ == "__main__":
    main()
